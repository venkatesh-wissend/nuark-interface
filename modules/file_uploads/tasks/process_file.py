import csv
from openpyxl import load_workbook
from modules.file_uploads.models.upload_data import UploadData
from modules.file_uploads.models.upload_log import UploadLog

def is_empty_row(row_dict):
    return all(v is None or str(v).strip() == "" for v in row_dict.values())

def process_file_task(file_path, module=None):
    # 1. Create UploadLog
    filename = file_path.split("/")[-1]
    log = UploadLog.objects.create(filename=filename, module=module)

    # 2. Process file
    global_row = 1
    if file_path.lower().endswith(".csv"):
        process_csv(file_path, global_row, log, module)
    else:
        process_excel(file_path, global_row, log, module)

    return log.id  # return the ID of the log

def process_csv(file_path, start_row, log, module):
    batch = []
    batch_size = 5000
    row_counter = start_row

    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for row in reader:
            if is_empty_row(row):
                continue

            batch.append(row)

            if len(batch) >= batch_size:
                insert_batch(batch, row_counter, log, module)
                row_counter += len(batch)
                batch = []

    if batch:
        insert_batch(batch, row_counter, log, module)

def process_excel(file_path, start_row, log, module):
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    headers = next(rows)
    batch = []
    batch_size = 5000
    row_counter = start_row

    for row in rows:
        row_dict = dict(zip(headers, row))
        if is_empty_row(row_dict):
            continue

        batch.append(row_dict)
        if len(batch) >= batch_size:
            insert_batch(batch, row_counter, log, module)
            row_counter += len(batch)
            batch = []

    if batch:
        insert_batch(batch, row_counter, log, module)

    wb.close()

def insert_batch(records, start_idx, log, module):
    objs = []
    for idx, record in enumerate(records):
        objs.append(UploadData(
            row_id=start_idx + idx,
            data=record,
            upload_log=log,  # set the foreign key here
            module=module
        ))

    UploadData.objects.bulk_create(objs)

